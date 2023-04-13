from docxtpl import DocxTemplate
from datetime import datetime, timedelta
from openpyxl.reader.excel import load_workbook
from .constants import people


now = datetime.now()
periodic_number = datetime.strftime(now, "%d%m%Y")   # номер протокола периодических испытаний
data_pick = datetime.strftime(now - timedelta(days=14), "%d/%m/%Y")  # дата отбора изделия
start_data = datetime.strftime(now - timedelta(days=7), "%d/%m/%Y")  # дата начала испытаний
data_get_carry = datetime.strftime(now - timedelta(days=21), "%d/%m/%Y")  # дата приемосдаточных испытаний
end_data = start_data  # дата окончания испытаний
data_exp_device = datetime.strftime(
    (now - timedelta(days=7)).replace(year=now.year + 1), "%d/%m/%Y")  # дата окончания действия испытаний
result_number = datetime.strftime(now, "%d%m%Y")  # номер акта результата периодических испытаний
device_number = 666  # Зав номер прибора


def documents(
        device,  # прибор
        periodic_number=periodic_number,   # номер протокола периодических испытаний
        data_pick=data_pick,  # дата отбора изделия
        start_data=start_data,  # дата начала испытаний
        end_data=end_data,  # дата окончания испытаний
        data_exp_device=data_exp_device,  # дата окончания действия испытаний
        result_number=result_number,  # номер акта результата периодических испытаний
        device_number=device_number,  # Зав номер прибора
):
    context = {
        'number_protocol_periodic': periodic_number,  # номер протокола периодических испытаний
        'number_protocol_get_carry': periodic_number,  # номер протокола приемосдаточных испытаний
        'device_tu': device.tu,  # ТУ прибора
        'device_iclg': device.iclg,  # ИЦЛГ прибора
        'device_number': device_number,  # Зав номер прибора
        'device_name': device.name,  # наименование прибора
        'device_fullname': device.fullname,  # полное наименование прибора
        'name_ci': people['ЦИ'].name,  # представитель ЦИ
        'name_iil': people['ИИЛ'].name,  # представитель ИИЛ
        'name_otk': people['ОТК'].name,  # представитель ОТК
        'test_meter': device.test_devices,  # тестовое оборудование
        'measure_meter': device.measure_devices,  # измерительное оборудование
        'number_result_periodic': result_number,  # номер акта результата периодических испытаний
        'start_data': start_data,  # дата начала испытаний
        'end_data': end_data,  # дата окончания испытаний
        'data_exp_device': data_exp_device,  # дата окончания действия испытаний
        'author_job': people['я'].job,  # должность кто составил документ
        'author_name': people['я'].name,  # имя кто составил документ
        'name_header': people['гл инженер'].name,  # имя в шапке документа
        'job_header': people['гл инженер'].job,  # должность в шапке документа
        'data_pick': data_pick,  # дата отбора изделия
        'data_get_carry': data_get_carry,    # дата премосдаточных испытаний
    }
    template = DocxTemplate(f"templates/{device.latin_name}.docx")
    template.render(context)
    return template.save(
        f'{device.name.replace("/", "-")}_{device_number}'
        f'{datetime.strftime(now, "%d_%m_%Y")}.docx'
    )

def exel_logger(data, serial, block):
    name = 'templates/doc_logger.xlsx'
    file = load_workbook(name)
    file_page = file['Лист1']
    max_row = file_page.max_row
    last_doc_number = file_page[f'A{max_row}'].value
    doc_number = int(last_doc_number) + 1
    file_page.append([str(doc_number), str(data), str(serial), str(block)])
    file.save(name)
    file.close()
    return doc_number


def isint(s):
    try:
        int(s)
        return True
    except ValueError:
        return False